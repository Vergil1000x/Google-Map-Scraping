import asyncio
import aiohttp
import openpyxl
from bs4 import BeautifulSoup
from arsenic import get_session, browsers, services

website_urls = []

file_path = r"C:\Users\Vergil1000\Downloads\AesCliMal.xlsx"
workbook = openpyxl.load_workbook(file_path)
worksheet = workbook.active

for row in worksheet.iter_rows(
    min_row=1, max_col=1, max_row=worksheet.max_row, values_only=True
):
    cell_value = row[0]
    if (
        cell_value is not None
        and isinstance(cell_value, str)
        and cell_value.startswith("http")
    ):
        website_urls.append(cell_value)

workbook.close()


file_path = r"C:\Users\Vergil1000\Downloads\testX.xlsx"
workbook = openpyxl.load_workbook(file_path)
worksheet = workbook.active


async def astre(url, limit):
    service = services.Chromedriver(
        binary=r"C:\Users\Vergil1000\Downloads\chromedriver-win64\chromedriver-win64\chromedriver.exe"
    )
    browser = browsers.Chrome()
    browser.capabilities = {
        "goog:chromeOptions": {
            "args": [
                "--headless",
                "--disable-gpu",
                "--no-sandbox",
                "--disable-dev-shm-usage",
            ]
        }
    }
    async with limit:
        async with get_session(service, browser) as session:
            await session.get(url)
            await asyncio.sleep(2)
            html = await session.get_page_source()
            soup = BeautifulSoup(html, "html.parser")
            string_list = []
            title = soup.title.string.strip()
            print(f"Title of {url}: {title}")

            element = soup.find(class_="DUwDvf lfPIob")
            if element:
                text = element.get_text(strip=True)
                string_list.append(text)
                print(text)
            else:
                print("Element with class 'DUwDvf lfPIob' not found.")
                string_list.append("Not Available")

            element = soup.find(class_="F7nice")
            if element:
                text = element.get_text(strip=True)
                string_list.append(text)
                print(text)
            else:
                print("Element with class 'F7nice' not found.")
                string_list.append("Not Available")

            element = soup.find(attrs={"data-tooltip": "Copy address"})
            if element:
                text = element.get_text(strip=True)
                string_list.append(text)
                print(text)
            else:
                print("Element with /data-tooltip = Copy address/ not found.")
                string_list.append("Not Available")

            element = soup.find("a", {"data-tooltip": "Open booking link"})
            if element and "href" in element.attrs:
                href_value = element["href"]
                string_list.append(href_value)
                print("Href attribute:", href_value)
            else:
                print("Element not found or doesn't have an href attribute.")
                string_list.append("Not Available")

            element = soup.find("a", {"data-tooltip": "Open website"})
            if element and "href" in element.attrs:
                href_value = element["href"]
                string_list.append(href_value)
                print("Href attribute:", href_value)
            else:
                print("Element not found or doesn't have an href attribute.")
                string_list.append("Not Available")

            element = soup.find(attrs={"data-tooltip": "Copy phone number"})
            if element:
                text = element.get_text(strip=True)
                string_list.append(text)
                print(text)
            else:
                print("Element with /data-tooltip = Copy phone number/ not found.")
                string_list.append("Not Available")

            print(f"Title of {url}: {title}")
            for string in string_list:
                print(string)
            worksheet.append(string_list)


async def main():
    limit = asyncio.Semaphore(5)
    tasks = [astre(url, limit) for url in website_urls]
    await asyncio.gather(*tasks)
    workbook.save(file_path)
    workbook.close()


if __name__ == "__main__":
    asyncio.run(main())
